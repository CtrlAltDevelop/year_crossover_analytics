import re
from datetime import timedelta
from enum import Enum
from functools import lru_cache
from pathlib import Path
from typing import Tuple, Dict, Any, List, Optional

import numpy as np
import pandas as pd
from PySide6.QtCore import SignalInstance
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font
from pandas.core.tools import datetimes

from source.core.common.report_class import ReportClass


class TimeFrame(Enum):
    M1 = 1  # 1 minute
    M2 = 2  # 2 minutes
    M3 = 3  # 3 minutes
    M4 = 4  # 4 minutes
    M5 = 5  # 5 minutes
    M6 = 6  # 6 minutes
    M10 = 10  # 10 minutes
    M12 = 12  # 12 minutes
    M15 = 15  # 15 minutes
    M20 = 20  # 20 minutes
    M30 = 30  # 30 minutes
    H1 = 16385  # 1 hour
    H2 = 16386  # 2 hours
    H3 = 16387  # 3 hours
    H4 = 16388  # 4 hours
    H6 = 16390  # 6 hours
    H8 = 16392  # 8 hours
    H12 = 16396  # 12 hours
    D1 = 16408  # 1 day
    W1 = 32769  # 1 week
    MN1 = 49153  # 1 month


class Mt5Report(ReportClass):
    contract_size = 0
    sessions = []
    valid_types = {'buy': ['buy', 'buy limit'], 'sell': ['sell', 'sell limit']}

    def set_contract_size(self, contract_size: int):
        self.contract_size = contract_size

    def update_sessions(self, day_list: List):
        self.sessions = day_list

    # def get_file_via_dialog(self, base_path: Path = Path.cwd()) -> Tuple[str, Tuple[pd.DataFrame, Dict[str, Any]]]:
    #     """
    #     Opens a file dialog for the user to select a CSV file and processes it.
    #
    #     :param base_path: The initial directory for the file dialog. Defaults to the current working directory.
    #     :return: A tuple containing the selected file path as a string and the processed pandas DataFrame.
    #     :raises FileNotFoundError: If no file is selected.
    #     """
    #     title = "Select EXCEL file containing MT5 Trade Report"
    #     filetypes = [("EXCEL files", "*.xlsx")]
    #     return self._get_file_via_dialog(title, filetypes, base_path)

    def _analyzed_data_with_gui(self, file: Path, connect: Optional[Path], progress: SignalInstance) -> Tuple[
        pd.DataFrame, Dict[str, Any]]:
        report_df, orders_df, deals_df, styles, merged_cells = self._read_excel_files(file)
        deal_connect = pd.read_csv(connect, index_col=0).to_dict('dict')['0'] if connect else None
        return (self._merge_order_and_deals_with_gui(orders_df, deals_df, progress, deal_connect),
                dict(df=report_df, styles=styles, merged_cells=merged_cells))

    @staticmethod
    def _extract_styles(ws, orders_start_idx):
        """
        Efficiently extract styles from a worksheet with caching for duplicates.
        """
        styles = {}

        @lru_cache(maxsize=None)
        def get_font(name, size, bold, italic, color):
            return Font(name=name, size=size, bold=bold, italic=italic, color=color)

        @lru_cache(maxsize=None)
        def get_fill(start_color, end_color, fill_type):
            return PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)

        @lru_cache(maxsize=None)
        def get_border(left, right, top, bottom):
            return Border(left=left, right=right, top=top, bottom=bottom)

        @lru_cache(maxsize=None)
        def get_alignment(horizontal, vertical, wrap_text):
            return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)

        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            if r_idx == orders_start_idx:
                break
            for c_idx, cell in enumerate(row, start=1):
                fnt = cell.font or Font()
                fill = cell.fill or PatternFill()
                border = cell.border or Border()
                align = cell.alignment or Alignment()

                styles[(r_idx, c_idx)] = {
                    "font": get_font(fnt.name, fnt.size, fnt.bold, fnt.italic, fnt.color),
                    "fill": get_fill(fill.start_color, fill.end_color, fill.fill_type),
                    "border": get_border(border.left, border.right, border.top, border.bottom),
                    "alignment": get_alignment(align.horizontal, align.vertical, align.wrap_text),
                    "number_format": cell.number_format,
                }

        return styles

    def _read_excel_files(self, path: Path) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict, List]:
        wb = load_workbook(filename=path, read_only=False, keep_vba=True)
        if "Deals and Orders" in wb.sheetnames:
            ws = wb["Deals and Orders"]  # Select "Deals and Orders" sheet
        else:
            ws = wb.active

        # Convert the worksheet to a DataFrame
        data = ws.values
        df = pd.DataFrame(data)

        # Find the indices for "Orders" and "Deals"
        orders_start_idx = df[df.iloc[:, 0] == "Orders"].index[0]
        deals_start_idx = df[df.iloc[:, 0] == "Deals"].index[0]

        # Extract the data from the beginning of the file to the "Orders" index
        result_df = df.iloc[:orders_start_idx].reset_index(drop=True)

        # Extract the "Orders" data
        orders_df = df.iloc[orders_start_idx + 2:deals_start_idx].reset_index(drop=True)
        orders_df.columns = df.iloc[orders_start_idx + 1].tolist()
        orders_df = orders_df.dropna(axis=1, how="all")
        orders_df = orders_df[orders_df['State'] == 'filled']

        deals_df = df.iloc[deals_start_idx + 2: -1].reset_index(drop=True)
        deals_df.columns = df.iloc[deals_start_idx + 1].tolist()

        orders_df = orders_df[['Open Time', 'Type', 'S / L', 'T / P', 'Time']].dropna(subset=['S / L', 'T / P'])
        orders_df.rename(columns={'Type': 'Type_y', 'Open Time': 'OpenTime', 'S / L': 'SL', 'T / P': 'TP'},
                         inplace=True)
        deals_df = deals_df[deals_df['Type'].isin(self.valid_types.keys())][[
            'Deal', 'Time', 'Type', 'Direction', 'Volume', 'Price', 'Commission', 'Swap', 'Profit', 'Balance']]
        deals_df.rename(columns={'Type': 'Type_x'}, inplace=True)

        return result_df, orders_df, deals_df, self._extract_styles(ws, orders_start_idx), ws.merged_cells.ranges

    def _merge_order_and_deals_with_gui(self, orders_df: pd.DataFrame, deals_df: pd.DataFrame, progress: SignalInstance,
                                        connect: Optional[Dict[int, int]]) -> pd.DataFrame:
        merged_data = []
        for _, deal_row in deals_df.iterrows():
            matching_orders = orders_df[orders_df['Time'] == deal_row['Time']]
            merged_row = deal_row.to_dict()

            valid_order = None
            if not matching_orders.empty:
                for _, order_row in matching_orders.iterrows():
                    if order_row['Type_y'] in self.valid_types[deal_row['Type_x']]:
                        valid_order = order_row
                        break
            if valid_order is not None:
                merged_row.update(valid_order.to_dict())
            merged_data.append(merged_row)
        merged_df = pd.DataFrame(merged_data)
        merged_df.set_index(keys='Deal', drop=True, inplace=True)
        merged_df['Volume'] = merged_df['Volume'].dropna().apply(self._convert_to_float)

        if connect:
            results_df = self.merge_with_connect_file(merged_df, progress, connect)
        else:
            results_df = self.merge_by_calculation(merged_df, progress)

        # results_df.to_csv('merged_orders_report.csv', index=False)
        return results_df

    @staticmethod
    def merge_with_connect_file(merged_df: pd.DataFrame, progress: SignalInstance,
                                connect: Dict[int, int]):
        open_rows = merged_df.loc[np.array(list(connect.keys()))]
        close_rows = merged_df.loc[np.array(list(connect.values()))]

        results = []
        for i in range(len(open_rows)):
            row = open_rows.iloc[i]
            opp_row = close_rows.iloc[i]
            results.append({
                'OrderTime': row['OpenTime'],
                'OpenTime': row['Time'],
                'Type': row['Type_y'].lower().replace('_', ''),
                'Lot': row['Volume'],
                'OpenPrice': row['Price'],
                'S/L': row['SL'],
                'T/P': row['TP'],
                'CloseTime': opp_row['Time'],
                'ClosePrice': opp_row['Price'],
                'Commission': row['Commission'] + opp_row['Commission'],
                'Swap': opp_row['Swap'],
                'Profit': opp_row['Profit'],
                'Balance': opp_row['Balance']
            })
            progress.emit(i * 100 / len(open_rows))

        results_df = pd.DataFrame(results)
        results_df[['OpenTime', 'OrderTime', 'CloseTime']] = results_df[['OpenTime', 'OrderTime', 'CloseTime']].apply(
            pd.to_datetime)
        results_df.loc[results_df['Type'].isin(['sell', 'buy']), 'OrderTime'] = np.nan
        return results_df

    def merge_by_calculation(self, merged_df: pd.DataFrame, progress: SignalInstance):
        in_deals = merged_df[merged_df['Direction'] == 'in'].sort_values(by='Time')
        opposite = merged_df[merged_df['Direction'] == 'out'].sort_values(by='Time')
        results = []
        indexes = set()
        current = 0
        for index, row in in_deals.iterrows():
            current += 1
            found = False
            matches = opposite[(opposite['Time'] >= row['Time']) & (opposite['Type_x'] != row['Type_x'])]
            for CIdx, opp_row in matches.iterrows():
                if CIdx in indexes:
                    continue
                if opp_row['Volume'] == row['Volume']:
                    profit = row['Volume'] * self.contract_size * (opp_row['Price'] - row['Price']) \
                        if row['Type_x'] == 'buy' \
                        else row['Volume'] * self.contract_size * (row['Price'] - opp_row['Price'])

                    if (opp_row['Price'] in [row['SL'], row['TP']]) or (opp_row['Profit'] == round(profit, 2)):
                        results.append({
                            'OrderTime': row['OpenTime'],
                            'OpenTime': row['Time'],
                            'Type': row['Type_y'].lower().replace('_', ''),
                            'Lot': row['Volume'],
                            'OpenPrice': row['Price'],
                            'S/L': row['SL'],
                            'T/P': row['TP'],
                            'CloseTime': opp_row['Time'],
                            'ClosePrice': opp_row['Price'],
                            'Commission': row['Commission'] + opp_row['Commission'],
                            'Swap': opp_row['Swap'],
                            'Profit': opp_row['Profit'],
                            'Balance': opp_row['Balance']
                        })
                        indexes.add(CIdx)
                        found = True
                        break
            if not found:
                for CIdx, opp_row in matches.iterrows():
                    if CIdx not in indexes and opp_row['Volume'] == row['Volume']:
                        results.append({
                            'OrderTime': row['OpenTime'],
                            'OpenTime': row['Time'],
                            'Type': row['Type_y'].lower().replace('_', ''),
                            'Lot': row['Volume'],
                            'OpenPrice': row['Price'],
                            'S/L': row['SL'],
                            'T/P': row['TP'],
                            'CloseTime': opp_row['Time'],
                            'ClosePrice': opp_row['Price'],
                            'Commission': row['Commission'] + opp_row['Commission'],
                            'Swap': opp_row['Swap'],
                            'Profit': opp_row['Profit'],
                            'Balance': opp_row['Balance']
                        })
                        indexes.add(CIdx)
                        break
            progress.emit(current * 100 / len(in_deals))

        results_df = pd.DataFrame(results)
        results_df[['OpenTime', 'OrderTime', 'CloseTime']] = results_df[['OpenTime', 'OrderTime', 'CloseTime']].apply(
            pd.to_datetime)
        results_df.loc[results_df['Type'].isin(['sell', 'buy']), 'OrderTime'] = np.nan
        return results_df

    @lru_cache(maxsize=1024)
    def _get_session_minutes_for_day(self, weekday: int) -> List[Tuple[int, int]]:
        return [(int(sess["Open"]), int(sess["Close"])) for sess in self.sessions[weekday]]

    def calculate_duration(self, open_time, close_time):
        if pd.isna(open_time) or pd.isna(close_time):
            return 0

        open_time, close_time = sorted([open_time, close_time])
        total_minutes = 0
        current = open_time

        while current.date() <= close_time.date():
            sessions = self._get_session_minutes_for_day((current.weekday() + 1) % 7)
            t_start = current.hour * 60 + current.minute + current.second / 60 if current.date() == open_time.date() else 0
            t_end = close_time.hour * 60 + close_time.minute + close_time.second / 60 if current.date() == close_time.date() else 1440

            for open_min, close_min in sessions:
                total_minutes += max(0, min(t_end, close_min) - max(t_start, open_min))
            current += timedelta(days=1)

        return round(total_minutes)

    def add_additional_data(self, df, commission: float, point: float, margin: float, calculate_commission: bool,
                            risk_value: float, balance: Optional[float] = None,
                            margin_file_path: Optional[Path] = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        df = df.copy()
        earliest_close_idx = df['CloseTime'].idxmin()
        df.at[earliest_close_idx, 'Balance'] -= df.at[earliest_close_idx, 'Swap'] + df.at[earliest_close_idx, 'Profit']
        df = df.sort_values(by='OpenTime', ascending=True)

        # Commission calculation (vectorized)
        if not calculate_commission:
            df['Commission'] = -np.round(df['Lot'] * commission, 2)

        # Margin and Equity calculation
        if margin_file_path:
            # Create key_merge correctly using per-row string conversion
            df['key_merge'] = df['OpenTime'].astype(str) + df['OpenPrice'].astype(str) + df['Lot'].astype(str)
            merge_df = df.merge(self._get_margin_via_path(margin_file_path), on='key_merge', how='left')
            df['Equity'] = merge_df['Margin'] + merge_df['FreeMargin']
            df['Margin ($)'] = merge_df['Margin']
            df['Margin_Balance'] = merge_df['Margin_Balance']
            df.drop(columns=['key_merge'], inplace=True)
        else:
            df['Equity'] = df['Balance']
            df['Margin ($)'] = np.round(df['Lot'] * margin, 2)
            df['Margin_Balance'] = df['Balance']

        # Margin Level calculation (vectorized)
        df['Margin Level (%)'] = np.round(df['Equity'] / df['Margin ($)'] * 100, 2)
        df['Margin Level (%) WorstCase'] = 0.0

        # Balance and Final Profit
        if balance is None:
            balance = df['Balance'].iloc[0]
        df['Final Profit'] = df['Swap'] + df['Profit'] + df['Commission']
        df['Balance'] = balance + np.cumsum(np.concatenate([[0], df['Final Profit'].iloc[:-1]]))

        df['Duration (Minutes)'] = df.apply(lambda r: self.calculate_duration(r['OpenTime'], r['CloseTime']), axis=1)
        df['Duration'] = pd.to_timedelta(df['Duration (Minutes)'], unit='m')

        df['CloseTimeShift'] = df['CloseTime'].shift(1)
        df['Time Between (Minutes)'] = df.apply(
            lambda r: self.calculate_duration(r['OpenTime'], r['CloseTimeShift']) if pd.notna(
                r['CloseTimeShift']) else pd.NA, axis=1)
        df['Time Between'] = pd.to_timedelta(df['Time Between (Minutes)'], unit='m')
        df.drop(columns=['CloseTimeShift'], inplace=True)

        # # Duration calculation (vectorized)
        # df['Duration'] = (df['CloseTime'] - df['OpenTime']).dt.round('s')
        # df['Duration (Minutes)'] = (df['Duration'].dt.total_seconds() / 60).round(0).astype(int)
        #
        # # Vectorized Time Between calculation
        # df['Time Between'] = df['OpenTime'] - df['CloseTime'].shift(1)
        # df['Time Between'] = df['Time Between'].where(df['Time Between'] > pd.Timedelta(0), pd.NaT)
        # df['Time Between (Minutes)'] = (df['Time Between'].dt.total_seconds() / 60).round(0).astype('Int64') + 1

        # Add Order column
        df.insert(0, 'Order', np.arange(1, len(df) + 1))

        # Concurrent trades analysis
        max_concurrent_buy = 0
        max_concurrent_sell = 0
        df['InSameTime'] = ''
        df['OppositeOrder_InSameTime'] = ''

        for i, row in df.iterrows():
            # Efficiently find concurrent trades using boolean indexing
            concurrent_mask = (df['OpenTime'] <= row['OpenTime']) & (df['CloseTime'] > row['OpenTime']) & (
                    df['Order'] != row['Order'])
            concurrent_trades = df[concurrent_mask]

            if concurrent_trades.empty:
                margin_value = np.round(row['Lot'] * margin, 2)
                df.at[i, 'Margin Level (%) WorstCase'] = df.at[i, 'Margin Level (%)']
            else:
                buy_trades = concurrent_trades[concurrent_trades['Type'].isin(['buy', 'buy limit'])]
                sell_trades = concurrent_trades[concurrent_trades['Type'].isin(['sell', 'sell limit'])]
                margin_value = np.round(abs(buy_trades['Lot'].sum() - sell_trades['Lot'].sum()) * margin, 2)

                # Update concurrent counts
                num_buy = len(buy_trades)
                num_sell = len(sell_trades)
                max_concurrent_buy = max(max_concurrent_buy, num_buy)
                max_concurrent_sell = max(max_concurrent_sell, num_sell)

                # Assign concurrent order IDs
                concurrent_order_ids = ', '.join(map(str, concurrent_trades['Order'].tolist()))
                df.at[i, 'InSameTime'] = concurrent_order_ids

                # Check for opposite orders
                if num_buy > 0 and num_sell > 0:
                    df.at[i, 'OppositeOrder_InSameTime'] = concurrent_order_ids

                # Worst-case margin level
                df.at[i, 'Margin Level (%) WorstCase'] = np.round(
                    (df.at[i, 'Margin_Balance'] - (risk_value * len(concurrent_trades))) / df.at[i, 'Margin ($)'] * 100,
                    2)

            if not margin_file_path:
                df.at[i, 'Margin ($)'] = margin_value

        df['Margin Level (%)'] = np.round(df['Equity'] / df['Margin ($)'] * 100, 2)
        df['Open_SL'] = (df['OpenPrice'] - df['S/L']) / point
        df['Open_Close'] = (df['OpenPrice'] - df['ClosePrice']) / point
        df.drop(columns=['Margin_Balance'], inplace=True)
        return df, {'buy': max_concurrent_buy, 'sell': max_concurrent_sell}

    @staticmethod
    @lru_cache(maxsize=None)
    def _convert_to_float(value):
        if isinstance(value, str) and 'M' in value:
            return float(value.replace('M', '').strip()) * 1000000
        if isinstance(value, str) and 'K' in value:
            return float(value.replace('K', '').strip()) * 1000
        return float(value)

    @staticmethod
    def get_modify_via_path(path: Path) -> Dict[datetimes, Dict[str, Any]]:
        result = {}
        for _, row in pd.read_csv(path).iterrows():
            result[pd.to_datetime(row['openDate'])] = \
                {'Time': pd.to_datetime(row['ModifyDate']), 'S/L': float(row['newSL'])}
        return result

    @staticmethod
    def _get_rev_name(text: str) -> str:
        match = re.search(r'Rev\S*', text)
        return match.group(0) if match else text

    def get_symbol_name(self, path) -> Dict[str, Any]:
        df = pd.read_excel(path)
        name = self._get_rev_name(df.iloc[2].dropna()['Unnamed: 3'])
        symbol = df.iloc[3].dropna()['Unnamed: 3']
        _time = df.iloc[4].dropna()['Unnamed: 3'].split(' ', 1)
        reward_to_risk, timeframe = None, None

        for index, row in df['Unnamed: 3'].items():
            if str(row).startswith(('_Inp_RewardToRisk', 'Inp_RewardToRisk')):
                reward_to_risk = float(row.replace('_Inp_RewardToRisk=', '').replace('Inp_RewardToRisk=', ''))
            if str(row).startswith(('_Inp_TradeTimeFrame', 'Inp_TradeTimeFrame')):
                timeframe = TimeFrame(
                    int(row.replace('_Inp_TradeTimeFrame=', '').replace('Inp_TradeTimeFrame=', ''))).name
            if (reward_to_risk and timeframe) or index == 100:
                break
        if not timeframe:
            timeframe = _time[0]
        start_time, end_time = _time[1].strip("()").split(" - ")
        return {'name': name, 'symbol': symbol, 'tf': timeframe, 'start_time': pd.to_datetime(start_time),
                'end_time': pd.to_datetime(end_time), 'reward_to_risk': reward_to_risk}

    def _get_margin_via_path(self, path: Path):
        df = pd.read_csv(path)
        df['OpenTime'] = df['OpenTime'].apply(pd.to_datetime)
        df['Lot'] = df['Lot'].apply(self._convert_to_float)
        df['key_merge'] = df['OpenTime'].astype(str) + df['OpenPrice'].astype(str) + df['Lot'].astype(str)
        df.rename(columns={'Balance': 'Margin_Balance'}, inplace=True)
        return df[['key_merge', 'Margin', 'FreeMargin', 'Margin_Balance']]
